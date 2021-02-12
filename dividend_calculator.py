import openpyxl
import sys

#Constants
TRANSACTION_STR_FOR_DIVIDEND = "ACH"
TRANSACTION_REMARK_COL = 5
DEPOSIT_AMT_COL = 7

passed_arguments = len(sys.argv)
if passed_arguments < 2:
  print("Input file name as command line argument")
  sys.exit()

print("passed_arguments:{}".format(passed_arguments))
print("\nName of file:", sys.argv[1])

#load the input bank statement
wb = openpyxl.load_workbook(sys.argv[1]) 
total_sheets = len(wb.sheetnames)

#total dividend received in a year
total_dividend = 0

#Contains total dividend given by company in a year
totCompanyDividend = {}

sheetNo = 0
#scroll through all the sheets in bank statement excel sheet
while (sheetNo < total_sheets):
  ws = wb.worksheets[sheetNo]
  
  rowNo = 1
  print("SHEET NO:{}".format(sheetNo+1))

  #Read till last row in the worksheet
  while rowNo <= ws.max_row:
    #Fetch the transaction remark from row and column
    txn_remark = ws.cell(row=rowNo,column=TRANSACTION_REMARK_COL).value
 
    #print("txn_remark:{}".format(txn_remark))
    #Handle error cases
    if (txn_remark is None) or ((txn_remark.find("Withdrawal Amount") != -1) or (txn_remark.find(TRANSACTION_STR_FOR_DIVIDEND) == -1)):
      rowNo = rowNo + 1
      continue

    #Fetch the dividend amount from row and column 
    dividend_amt = ws.cell(row=rowNo,column=DEPOSIT_AMT_COL).value

    txn_remark = txn_remark.replace("\n", "")
    print("   ** {} : {}".format(txn_remark, dividend_amt))
  
    #Add dividend amount in total_dividend
    total_dividend = total_dividend + float(dividend_amt)
    rowNo = rowNo + 1        
      
  sheetNo = sheetNo + 1

print("\n\n\n***** Total dividend received:{}\n\n".format(total_dividend))
